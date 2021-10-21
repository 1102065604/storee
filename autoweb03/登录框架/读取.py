from openpyxl import load_workbook

class Write_excel(object):
    # '修改excel数据'
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb['login']  # 激活sheet

    def write1(self, row_n1, col_n1, value1):
       #  写入数据，如(2,3,"hello"),第二行第三列写入数据"hello"
        self.ws.cell(row_n1, col_n1, value1)
        self.wb.save(self.filename)

class Write_excel1(object):
    # '修改excel数据'
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb['error']  # 激活sheet

    def write1(self, row_n1, col_n1, value1):
       #  写入数据，如(2,3,"hello"),第二行第三列写入数据"hello"
        self.ws.cell(row_n1, col_n1, value1)
        self.wb.save(self.filename)








