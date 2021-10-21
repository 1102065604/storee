import ntpath

from selenium import webdriver
from unittest import TestCase
from ddt import ddt
from ddt import data
from ddt import unpack
from 数据源 import InitPageData
from 实际操作 import LoginOpera
from duqu import Write_excel
import time
import xlrd
import xlwt
import openpyxl
import os
from xlutils.copy import copy
'''
    成功的用例
    失败的用例
'''
n=1
@ddt
class TestLogin(TestCase):
    #在每个测试用例执行前执行
    def setUp(self) -> None:
        self.driver = webdriver.Chrome()
        self.driver.get("http://localhost:8080/HKR")
        self.driver.maximize_window()

    # 在每个用例执行后执行
    def tearDown(self) -> None:
        self.driver.quit()


    @data(*InitPageData.da)
    # @unpack
    def testLoginSuccess(self, testdata):
        username = testdata["username"]
        password = testdata["password"]
        expect = testdata["expect"]

        time.sleep(2)
        loginopr = LoginOpera(self.driver)
        #  登陆的三项操作
        loginopr.login(username, password)

        result = loginopr.getSuccess_result()
        time.sleep(2)

        # wx=self.assertEqual(result,expect)

        #
        # wx=openpyxl.load_workbook(filename=r'C:\Users\LENOVO\Desktop\python\自动化\day03\HKR.xlsx')
        # sheetnames = wx.get_sheet_names()
        # table = wx.get_sheet_by_name(sheetnames[0])
        # table = wx.active

        # nrows = table.max_row  # 获得行数
        # ncolumns = table.max_column  # 获得行数

        global n
        table = Write_excel(r"D:\python\pythondaima\autoweb\autoweb03\框架\HKR.xlsx")
        if result == expect:
            table.write(n + 1, 4, 'pass')

        else:
            table.write(n + 1, 4, 'fail')
        n += 1
        self.assertEqual(result, expect)



    # @data(*InitPageData.login_error_data)
    # def testLoginError(self, testdata):
    #     username = testdata["username"]
    #     password = testdata["password"]
    #     expect = testdata["expect"]
    #     time.sleep(2)
    #     loginopr = LoginOpera(self.driver)
    #     #  登陆的三项操作
    #     loginopr.login(username, password)
    #
    #     result = loginopr.getError_result()
    #     time.sleep(2)
    #
    #     # 将测试结果回写到excel表里，xlwt
    #     self.assertEqual(result, expect)